import os
import openpyxl
from  tkinter import *
import tkinter as tk
from tkinter import filedialog

ventana = Tk()

ventana.title("Generador de Excel")

ventana.geometry("750x365")


marco= Frame(ventana,width=735,height=350)
marco.config(
    bg="white",
    bd=2,
)
marco.pack(padx=5 , pady=6)

marco.pack_propagate(FALSE)
texto=Label(marco,text="Expecifique una ruta: ")
texto.config(
    bg="white",
    font=("Arial",14)
)

datos=StringVar()
texto.pack(anchor=CENTER,side=LEFT)
campo_txt=Entry(marco)
campo_txt.config(
    width=65,font=("Arial",10),
    border=2,
    textvariable=datos
    
)

def abrirCarpeta():

    raiz=tk.Tk()
    raiz.withdraw()
    ruta_archivos=filedialog.askdirectory()
    datos.set(ruta_archivos)


import os
import openpyxl

def getdatos():
    ruta = str(datos.get())
    rutaAbsoluta = os.path.abspath(ruta)

    excel = openpyxl.Workbook()
    hoja = excel.active

    hoja.cell(row=1, column=1, value=ruta)
    contador = 2
    cantidad_archivos = 0
    for archivo in os.listdir(rutaAbsoluta):
        if os.path.isfile(os.path.join(rutaAbsoluta, archivo)):
            cantidad_archivos += 1

    hoja.cell(row=1, column=10, value="Nº de archivos: " + str(cantidad_archivos))


    for raiz, directorios, archivos in os.walk(rutaAbsoluta):
        for carpeta in directorios:
            ruta_completa = os.path.join(raiz, carpeta)
            cantidad_archivos = 0
            for archivo in os.listdir(ruta_completa):
                if os.path.isfile(os.path.join(ruta_completa, archivo)):
                    cantidad_archivos += 1
            hoja.cell(row=contador, column=1, value=carpeta)
            hoja.cell(row=contador, column=10, value="Nº de archivos: " + str(cantidad_archivos))
            contador += 1

    excel.save("Datos.xlsx")






campo_txt.pack(anchor=CENTER,side=LEFT)
btn=Button(marco)
btn.config(
  text="Buscar",
  height=2,
  width=8,
  cursor="hand2" ,
  bg="grey",
  fg="white",
  command=abrirCarpeta
  
)

btn.pack(side=LEFT,anchor=E)
btn_excel=Button(marco)
btn_excel.config(
  text="Generar Excel",
  height=2,
  width=12,
  cursor="hand2" ,
  bg="grey",
  fg="white",
  command=getdatos
)
btn_excel.place(x=375,y=215)
ventana.mainloop()
